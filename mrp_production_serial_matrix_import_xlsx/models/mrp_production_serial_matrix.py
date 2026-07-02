# Copyright 2026 ForgeFlow S.L. (https://www.forgeflow.com)
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl.html).

# flake8: noqa: B904

import base64
import io

import xlsxwriter

from odoo import _, fields, models
from odoo.exceptions import UserError
from odoo.tools.float_utils import float_is_zero

try:
    import openpyxl
except ImportError:
    openpyxl = None


class MrpProductionSerialMatrix(models.Model):
    _inherit = "mrp.production.serial.matrix"

    import_file = fields.Binary(string="Upload Excel File")
    import_filename = fields.Char(string="File Name")

    def action_download_template(self):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        worksheet = workbook.add_worksheet("Template")

        header_format = workbook.add_format(
            {"bold": True, "bg_color": "#f2f2f2", "border": 1}
        )

        unique_comp_headers = list(
            dict.fromkeys(self.line_ids.mapped("component_column_name"))
        )
        headers = ["Finished Product Serial Numbers"] + unique_comp_headers

        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
            worksheet.set_column(col_num, col_num, 25)

        matrix_data = {}

        for line in self.line_ids:
            finished_sn = line.finished_lot_name

            if finished_sn not in matrix_data:
                matrix_data[finished_sn] = {}

            matrix_data[finished_sn][line.component_column_name] = (
                line.component_lot_id.name or ""
            )

        row_num = 1
        for finished_sn, components_map in matrix_data.items():
            worksheet.write(row_num, 0, finished_sn)

            for col_idx, col_header in enumerate(unique_comp_headers, start=1):
                cell_value = components_map.get(col_header, "")
                worksheet.write(row_num, col_idx, cell_value)

            row_num += 1

        workbook.close()
        output.seek(0)

        attachment_id = self.env["ir.attachment"].create(
            {
                "name": f"{self.display_name} - {_('Serial Matrix Template')}",
                "datas": base64.encodebytes(output.getvalue()),
            }
        )
        output.close()
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment_id.id}",
            "target": "download",
        }

    def _xlsx_to_data_map(self, sheet):
        header = [cell.value for cell in sheet[1]]
        col_mapping = {name: i for i, name in enumerate(header)}
        data_map = {}
        finished_lot_names = []
        for row_x, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):  # noqa: B007
            if not row or not row[0]:
                continue
            finished_lot_name = str(row[0])
            if finished_lot_name not in finished_lot_names:
                finished_lot_names.append(finished_lot_name)
            for col_name, col_idx in col_mapping.items():
                if col_name == "Finished Product Serial Numbers" or not row[col_idx]:
                    continue
                data_map[(finished_lot_name, col_name)] = str(row[col_idx]).strip()
        return finished_lot_names, data_map

    def _get_or_create_finished_lots(self, lot_names):
        finished_lot_ids = self.env["stock.lot"]
        for name in lot_names:
            lot = self.env["stock.lot"].search(
                [("name", "=", name), ("product_id", "=", self.product_id.id)], limit=1
            )
            if not lot:
                lot = self.env["stock.lot"].create(
                    {
                        "name": name,
                        "product_id": self.product_id.id,
                        "company_id": self.company_id.id,
                    }
                )
            finished_lot_ids |= lot
        return finished_lot_ids

    def _get_component_templates(self):
        tracked_components = []
        for move in self.production_id.move_raw_ids:
            rounding = move.product_id.uom_id.rounding
            if float_is_zero(move.product_qty, precision_rounding=rounding):
                continue
            qty_per_unit = move.product_qty / self.production_id.product_qty
            if move.product_id.tracking == "serial":
                for i in range(1, int(qty_per_unit) + 1):
                    tracked_components.append(
                        {
                            "component_id": move.product_id.id,
                            "component_column_name": f"{move.product_id.display_name} "
                            f"({i})",
                            "lot_qty": 1,
                        }
                    )
            elif move.product_id.tracking == "lot" and self.include_lots:
                tracked_components.append(
                    {
                        "component_id": move.product_id.id,
                        "component_column_name": move.product_id.display_name,
                        "lot_qty": qty_per_unit,
                    }
                )
        return tracked_components

    def _prepare_matrix_lines_from_data(
        self, finished_lot_ids, component_templates, data_map
    ):
        new_lines_vals = []
        for lot in finished_lot_ids:
            for comp_template in component_templates:
                vals = comp_template.copy()
                vals.update({"finished_lot_id": lot.id, "finished_lot_name": lot.name})
                comp_lot_name = data_map.get((lot.name, vals["component_column_name"]))
                if comp_lot_name:
                    comp_lot = self.env["stock.lot"].search(
                        [
                            ("name", "=", comp_lot_name),
                            ("product_id", "=", vals["component_id"]),
                            ("company_id", "=", self.company_id.id),
                        ],
                        limit=1,
                    )
                    if not comp_lot:
                        raise UserError(
                            _(
                                "Component Lot '%(comp_lot_name)s' not found for "
                                "product '%(comp_col_name)s' (SN: %(lot_name)s)"
                            )
                            % {
                                "comp_lot_name": comp_lot_name,
                                "comp_col_name": vals["component_column_name"],
                                "lot_name": lot.name,
                            }
                        )
                    vals["component_lot_id"] = comp_lot.id
                new_lines_vals.append(vals)
        return new_lines_vals

    def action_import_template(self):
        self.ensure_one()
        if not self.import_file:
            raise UserError(_("Please upload an Excel file first."))
        if not openpyxl:
            raise UserError(_("The 'openpyxl' library is required for this action."))

        try:
            file_content = base64.b64decode(self.import_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Failed to read the file: %s") % e)

        finished_lot_names, data_map = self._xlsx_to_data_map(sheet)
        finished_lot_ids = self._get_or_create_finished_lots(finished_lot_names)
        component_templates = self._get_component_templates()
        new_lines_vals = self._prepare_matrix_lines_from_data(
            finished_lot_ids, component_templates, data_map
        )

        if new_lines_vals or finished_lot_ids:
            self.line_ids = [(5, 0, 0)]
            self.line_ids = [(0, 0, v) for v in new_lines_vals]
            self.finished_lot_ids = [(6, 0, finished_lot_ids.ids)]
        else:
            raise UserError(_("The imported file contains no valid data."))

        return {"type": "ir.actions.client", "tag": "reload"}
